from openpyxl import Workbook
filepath = "yamete.xlsx"
wb = Workbook()

ws = wb.create_sheet("Mysheet",0)

# ws['A1'] = 'ayam'

list_var = ['A','B','C','D','E']
list_column = ['NO','NAME','DESC','EMAIL','URL']
a = 1
# ws['A'+str(a)] = 'No'
# ws['B'+str(a)] = 'Name'
# ws['C'+str(a)] = 'Desc'

# Loop and find index and value from list
for idx,var in enumerate(list_var):
    ws[list_var[idx] + str(a)] = list_column[idx]

wb.save(filepath)