
from openpyxl import load_workbook
wb2 = load_workbook('hello world.xlsx')
print(wb2.sheetnames)